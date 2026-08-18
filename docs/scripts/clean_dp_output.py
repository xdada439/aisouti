#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
clean_dp_output.py — 对 DP 输出的题库做规则化清洗

规则：
  【删除】
    · 正确答案为空的行
    · 题型 = 填空题（量小，没法可靠转换）
    · 题干为空 / 过短 (<5 字) / 过长 (>250 字)
    · 多题挤一行（题干含 "（X）2." 之类答案括号 + 续题号）

  【修复】
    · 题型 vs 答案数量倒推：
        - 答案 1 字母 + 题型"多选题" → 改成"单选题"
        - 答案 ≥2 字母 + 题型"单选题" → 改成"多选题"
    · 题干末尾的答案括号 "（C）"/"(A)" → 删除
    · 题干前缀题号 "1.", "(2)", "第10题" → 删除
    · 答案归一：单选取第一字母、多选去重排序、判断题→A/B
"""

import sys
import io
import re

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

from openpyxl import load_workbook, Workbook

SRC = r'C:\Users\86150\Desktop\新建文件夹\测试.xlsx'
DST = r'C:\Users\86150\Desktop\新建文件夹\测试_clean.xlsx'

HEADER = ['题干', '题型', '选项A', '选项B', '选项C', '选项D',
          '选项E', '选项F', '选项G', '选项H', '正确答案']

# "（A）2." 这类 = 多题挤一行（答案括号紧跟下一题题号）
MULTI_Q_PAT = re.compile(r'[（(][A-HＡ-Ｈ]+[）)]\s*[\d０-９]+\s*[、．.]')
# 题干末尾的答案括号
TAIL_ANS_PAT = re.compile(r'\s*[（(][A-HＡ-Ｈ]+[）)]\s*[。．.]?\s*$')
# 题干前缀题号
LEAD_NO_PAT = re.compile(r'^\s*(?:第?\s*\d+\s*题?\s*[、．.]\s*|[（(]\s*\d+\s*[）)]\s*)')

stats = {
    'total': 0, 'kept': 0,
    'drop_empty_answer': 0, 'drop_blank_type': 0,
    'drop_no_stem': 0, 'drop_too_short': 0, 'drop_too_long': 0,
    'drop_multi_q': 0, 'drop_bad_answer': 0,
    'drop_no_options': 0, 'drop_stem_has_options': 0,
    'drop_multi_q_in_stem': 0,
    'fix_type_s2m': 0, 'fix_type_m2s': 0,
    'fix_tail_paren': 0, 'fix_lead_qno': 0,
    'fix_judge_ans': 0,
}

# 题干里夹着选项（"...A...B...C...D..." 4+ 个连续选项标记）
STEM_HAS_OPTS_PAT = re.compile(r'[^a-zA-Z]A[^a-zA-Z].{2,80}?B[^a-zA-Z].{2,80}?C[^a-zA-Z].{2,80}?D[^a-zA-Z]')

# 题干含多个问号（疑似多题挤一行）
MULTI_QUESTION_MARK_PAT = re.compile(r'[？?].*[？?]')

src_wb = load_workbook(SRC, read_only=True)
src_ws = src_wb.active

dst_wb = Workbook()
dst_ws = dst_wb.active
dst_ws.title = '题库'
dst_ws.append(HEADER)

for i, row in enumerate(src_ws.iter_rows(values_only=True), 1):
    if i == 1:
        continue  # skip header

    stats['total'] += 1
    cells = list(row) + [None] * (11 - len(row))
    cells = cells[:11]

    stem = str(cells[0]).strip() if cells[0] is not None else ''
    qtype = str(cells[1]).strip() if cells[1] is not None else ''
    options = [str(cells[2 + j]).strip() if cells[2 + j] is not None else '' for j in range(8)]
    answer_raw = str(cells[10]).strip() if cells[10] is not None else ''

    # ── 删除规则 ──
    if not answer_raw:
        stats['drop_empty_answer'] += 1
        continue
    if qtype == '填空题':
        stats['drop_blank_type'] += 1
        continue
    if not stem:
        stats['drop_no_stem'] += 1
        continue
    if len(stem) < 5:
        stats['drop_too_short'] += 1
        continue
    if len(stem) > 250:
        stats['drop_too_long'] += 1
        continue
    if MULTI_Q_PAT.search(stem):
        stats['drop_multi_q'] += 1
        continue
    # 题干里直接夹着 A B C D 选项 → 题型识别会失败，砍掉
    if STEM_HAS_OPTS_PAT.search(stem):
        stats['drop_stem_has_options'] += 1
        continue
    # 题干含多个问号 → 多道题挤在一起，砍掉
    if MULTI_QUESTION_MARK_PAT.search(stem):
        stats['drop_multi_q_in_stem'] += 1
        continue

    # ── 修复：题干末尾答案括号 ──
    new_stem = TAIL_ANS_PAT.sub('', stem).strip()
    if new_stem != stem:
        stats['fix_tail_paren'] += 1
        stem = new_stem

    # ── 修复：题干前缀题号 ──
    new_stem = LEAD_NO_PAT.sub('', stem).strip()
    if new_stem != stem:
        stats['fix_lead_qno'] += 1
        stem = new_stem

    # 二次检查长度
    if len(stem) < 5:
        stats['drop_too_short'] += 1
        continue

    # ── 修复：题型 vs 答案 ──
    ans_upper = answer_raw.upper().replace(' ', '').replace(',', '').replace('，', '').replace(';', '').replace('、', '')
    ans_letters = re.findall(r'[A-H]', ans_upper)

    final_answer = ''
    if qtype in ('单选题', '多选题'):
        # 没有字母 → 看是不是判断题误标
        if not ans_letters:
            stats['drop_bad_answer'] += 1
            continue
        if len(ans_letters) == 1 and qtype == '多选题':
            qtype = '单选题'
            stats['fix_type_m2s'] += 1
        elif len(ans_letters) >= 2 and qtype == '单选题':
            qtype = '多选题'
            stats['fix_type_s2m'] += 1

        if qtype == '单选题':
            final_answer = ans_letters[0]
        else:
            final_answer = ''.join(sorted(set(ans_letters)))

    elif qtype == '判断题':
        # 判断题：对/错 → A/B
        if ans_upper in ('对', '正确', 'T', '√', 'A', '是'):
            final_answer = 'A'
            if answer_raw != 'A':
                stats['fix_judge_ans'] += 1
        elif ans_upper in ('错', '错误', 'F', '×', 'B', '否'):
            final_answer = 'B'
            if answer_raw != 'B':
                stats['fix_judge_ans'] += 1
        elif ans_letters and ans_letters[0] in ('A', 'B'):
            final_answer = ans_letters[0]
        else:
            stats['drop_bad_answer'] += 1
            continue
    else:
        # 未知题型 → 当单选处理（按答案）
        if not ans_letters:
            stats['drop_bad_answer'] += 1
            continue
        qtype = '多选题' if len(ans_letters) >= 2 else '单选题'
        final_answer = ans_letters[0] if len(ans_letters) == 1 else ''.join(sorted(set(ans_letters)))

    # ── 选项数检查（单选/多选必须 ≥ 2 个选项；判断题不需要）──
    non_empty_options = [o for o in options if o.strip()]
    if qtype in ('单选题', '多选题') and len(non_empty_options) < 2:
        stats['drop_no_options'] += 1
        continue

    # ── 写出 ──
    out_row = [stem, qtype] + options + [final_answer]
    dst_ws.append(out_row)
    stats['kept'] += 1

dst_wb.save(DST)

print('━━━━━━━━━━ 清洗完成 ━━━━━━━━━━')
print(f'')
print(f'📊 总览:')
print(f'  原始行数: {stats["total"]}')
print(f'  保留行数: {stats["kept"]}  ({stats["kept"]*100/stats["total"]:.1f}%)')
print(f'  删除合计: {stats["total"] - stats["kept"]}')
print(f'')
print(f'🗑  删除明细:')
print(f'  · 空答案:                 {stats["drop_empty_answer"]:>5} 行')
print(f'  · 填空题(整体砍掉):       {stats["drop_blank_type"]:>5} 行')
print(f'  · 无题干:                 {stats["drop_no_stem"]:>5} 行')
print(f'  · 题干<5字:               {stats["drop_too_short"]:>5} 行')
print(f'  · 题干>250字:             {stats["drop_too_long"]:>5} 行')
print(f'  · 多题挤一行(答案括号+续号): {stats["drop_multi_q"]:>5} 行')
print(f'  · 题干里夹着选项 ABCD:    {stats["drop_stem_has_options"]:>5} 行')
print(f'  · 题干含多问号:           {stats["drop_multi_q_in_stem"]:>5} 行')
print(f'  · 选项数<2 (无法识别):    {stats["drop_no_options"]:>5} 行')
print(f'  · 答案不可解析:           {stats["drop_bad_answer"]:>5} 行')
print(f'')
print(f'🔧 修复明细:')
print(f'  · 题型 单→多 (按答案):  {stats["fix_type_s2m"]:>5} 行')
print(f'  · 题型 多→单 (按答案):  {stats["fix_type_m2s"]:>5} 行')
print(f'  · 题干末答案括号删除:   {stats["fix_tail_paren"]:>5} 行')
print(f'  · 题干前题号删除:       {stats["fix_lead_qno"]:>5} 行')
print(f'  · 判断题答案归一(对/错→A/B): {stats["fix_judge_ans"]:>5} 行')
print(f'')
print(f'📁 输出: {DST}')
