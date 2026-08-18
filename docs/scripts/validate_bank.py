#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
validate_bank.py — 题库 Excel 校验脚本

校验 DP / 人工 处理出来的题库 Excel 是否符合"AI 搜题助手"导入规范。
不通过的项必须修复，否则导入后命中率为 0%。

用法:
    python validate_bank.py <input.xlsx>
    python validate_bank.py <input.xlsx> --strict   # 严格模式：任何错误都退出
    python validate_bank.py <input.xlsx> --sample 10  # 抽样肉眼复核（默认 5）

输出:
    控制台打印每项检查的 ✓/✗ + 错误明细 + 抽样数据
    退出码: 0 = 全通过, 1 = 有错误（严格模式）, 2 = 有警告
"""

import sys
import io
import re
import random
import argparse
from collections import Counter, defaultdict

# Windows 终端默认 GBK 编码会把中文 emoji 报错，强制 UTF-8
if sys.platform == 'win32':
    try:
        sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
        sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')
    except Exception:
        pass

try:
    from openpyxl import load_workbook
except ImportError:
    sys.exit("缺少依赖: pip install openpyxl")


REQUIRED_HEADERS = [
    '题干', '题型', '选项A', '选项B', '选项C', '选项D',
    '选项E', '选项F', '选项G', '选项H', '正确答案'
]
TYPES_OK = {'单选题', '多选题', '判断题', '填空题'}


# ──────────────────────────────────────────
# 检查项
# ──────────────────────────────────────────

class Issue:
    """单个问题。level: 'error'（致命）/ 'warn'（建议修）"""
    def __init__(self, level, row, message, sample=''):
        self.level = level
        self.row = row
        self.message = message
        self.sample = sample

    def __str__(self):
        prefix = '✗' if self.level == 'error' else '⚠'
        s = f"  {prefix} 行{self.row}: {self.message}"
        if self.sample:
            s += f"\n      样本: {self.sample[:120]}"
        return s


def check_header(ws):
    issues = []
    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    if len(header) < 11:
        issues.append(Issue('error', 1, f'列数 {len(header)} < 11，缺列'))
        return issues, header
    for i, expected in enumerate(REQUIRED_HEADERS):
        actual = header[i] if i < len(header) else None
        if actual != expected:
            issues.append(Issue('error', 1, f'列{i+1} 名称错: 实际={actual!r}, 应为={expected!r}'))
    return issues, header


def check_rows(ws):
    """逐行检查规则"""
    issues = []
    type_dist = Counter()
    ans_format_bad = 0
    empty_stem = 0
    empty_ans = 0
    qno_prefix = 0
    section_prefix = 0
    multi_in_one = 0
    illegal_type = 0
    sub_q_marker = 0
    digits_only_stem = 0
    short_stem = 0
    too_long_stem = 0

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if len(row) < 11:
            issues.append(Issue('error', row_idx, '列数 < 11，数据可能损坏'))
            continue

        stem = (row[0] or '')
        qtype = (row[1] or '')
        opts = [str(x or '').strip() for x in row[2:10]]
        ans = (row[10] or '')

        # 跳过完全空行
        if not any([stem, qtype, ans] + opts):
            continue

        stem_s = str(stem).strip()
        qtype_s = str(qtype).strip()
        ans_s = str(ans).strip()

        # 1. 题干为空
        if not stem_s:
            empty_stem += 1
            issues.append(Issue('error', row_idx, '题干为空'))
            continue

        # 2. 题干太短（< 4 字符）— 可能是脏数据
        if len(stem_s) < 4:
            short_stem += 1
            issues.append(Issue('error', row_idx, f'题干过短（{len(stem_s)} 字）', stem_s))

        # 3. 题干全是数字/符号
        if re.match(r'^[\d\.\s\-:：mol/Lkgd℃°%]+$', stem_s):
            digits_only_stem += 1
            issues.append(Issue('error', row_idx, '题干疑似只有数字/单位，可能是选项被错误存为题干', stem_s))

        # 4. 题干含题号前缀
        if re.match(r'^\d{1,4}[、.,)）.]\s*', stem_s):
            qno_prefix += 1
            issues.append(Issue('warn', row_idx, '题干含题号前缀（应去掉）', stem_s[:50]))

        # 5. 题干含章节标签
        if re.search(r'(一|二|三|四|五|六|七|八|九|十|百)[、．]|第[一二三四五六七八九十\d]+[章节部分]|[Aa]\d?型题|单选题（|多选题（', stem_s):
            # 如果是简单的"A1型题"等前缀可能可以接受，但多见于错误污染
            if re.match(r'^[一二三四五六七八九十]+[、.]|^第[\d一二三四五六七八九十]+[章节]|^[A-Za-z]\d?型题', stem_s):
                section_prefix += 1
                issues.append(Issue('error', row_idx, '题干含章节标签（应去掉）', stem_s[:60]))

        # 6. 题干含多道题（出现 "？" 或 "：" 后又有新句子开头）
        # 简单启发：题干超过 200 字 + 含 2+ 个 "？" 或 ":" 末尾
        if stem_s.count('?') + stem_s.count('？') >= 2 and len(stem_s) > 100:
            multi_in_one += 1
            issues.append(Issue('error', row_idx, '题干含多个问号，疑似多道题挤一行', stem_s[:80]))
        # 题干太长（> 250 字）— 案例题之外不应该
        if len(stem_s) > 250:
            too_long_stem += 1
            issues.append(Issue('warn', row_idx, f'题干超过 250 字（{len(stem_s)} 字），可能多题挤一行或案例未精简', stem_s[:80]))

        # 7. 选项内容跨行包含"题干"特征（"是："、"应"、"包括"等结尾）
        for i, opt in enumerate(opts):
            if not opt:
                continue
            # 选项字符跨过 80 字且末尾是 "：" / "?" → 疑似选项串了下一道题题干
            if len(opt) > 80 and (opt.endswith('：') or opt.endswith(':') or opt.endswith('?') or opt.endswith('？')):
                issues.append(Issue('warn', row_idx, f'选项{chr(65+i)} 末尾像题干', opt[:60]))

        # 8. 题型
        if qtype_s not in TYPES_OK:
            illegal_type += 1
            issues.append(Issue('error', row_idx, f'题型不合法: {qtype_s!r}（应为单选题/多选题/判断题/填空题之一）'))
        else:
            type_dist[qtype_s] += 1

        # 9. 正确答案为空
        if not ans_s:
            empty_ans += 1
            issues.append(Issue('error', row_idx, '正确答案为空'))
        else:
            # 答案格式检查
            if qtype_s == '单选题':
                if not re.fullmatch(r'[A-Ha-h]', ans_s):
                    ans_format_bad += 1
                    issues.append(Issue('error', row_idx, f'单选题答案格式错: {ans_s!r}（应为单字母 A-H）'))
            elif qtype_s == '多选题':
                if not re.fullmatch(r'[A-Ha-h]{2,}', ans_s):
                    ans_format_bad += 1
                    issues.append(Issue('error', row_idx, f'多选题答案格式错: {ans_s!r}（应≥2 个字母连写，无分隔符）'))
                if '、' in ans_s or ',' in ans_s or ' ' in ans_s or '|' in ans_s:
                    ans_format_bad += 1
                    issues.append(Issue('error', row_idx, f'多选题答案含分隔符: {ans_s!r}（应连写如 ABCD）'))
            elif qtype_s == '判断题':
                if ans_s not in ('A', 'B', '对', '错', '正确', '错误'):
                    issues.append(Issue('warn', row_idx, f'判断题答案: {ans_s!r}（建议 A=对 / B=错）'))

        # 10. 选项数量与题型匹配
        non_empty_opts = sum(1 for o in opts if o)
        if qtype_s in ('单选题', '多选题') and non_empty_opts < 2:
            issues.append(Issue('error', row_idx, f'{qtype_s} 但选项数 = {non_empty_opts}（应 ≥ 2）'))
        if qtype_s == '填空题' and non_empty_opts > 0:
            issues.append(Issue('warn', row_idx, f'填空题但有 {non_empty_opts} 个选项（填空题应全空）'))

        # 11. 选项内容含 "A." / "A、" 前缀（应剥）
        for i, opt in enumerate(opts):
            if opt and re.match(r'^[A-Ha-h][.、．:：\)）]\s*', opt):
                issues.append(Issue('warn', row_idx, f'选项{chr(65+i)} 含字母前缀: {opt[:30]}（应去掉前缀）'))

    return issues, {
        'type_dist': type_dist,
        'empty_stem': empty_stem,
        'empty_ans': empty_ans,
        'ans_format_bad': ans_format_bad,
        'qno_prefix': qno_prefix,
        'section_prefix': section_prefix,
        'multi_in_one': multi_in_one,
        'illegal_type': illegal_type,
        'digits_only_stem': digits_only_stem,
        'short_stem': short_stem,
        'too_long_stem': too_long_stem,
    }


def check_duplicates(ws):
    """题干完全重复的行"""
    seen = defaultdict(list)
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        stem = str(row[0] or '').strip()
        if stem and len(stem) > 5:
            seen[stem].append(row_idx)
    dup_groups = {k: v for k, v in seen.items() if len(v) > 1}
    issues = []
    for stem, rows in list(dup_groups.items())[:10]:  # 只报前 10 组
        issues.append(Issue('warn', rows[0], f'重复题（共 {len(rows)} 行: {rows[:5]}...）', stem[:60]))
    return issues, len(dup_groups)


def sample_rows(ws, n=5):
    """随机抽 n 行展示"""
    all_rows = list(ws.iter_rows(min_row=2, values_only=True))
    if len(all_rows) <= n:
        return all_rows
    return random.sample(all_rows, n)


# ──────────────────────────────────────────
# 主入口
# ──────────────────────────────────────────

def main():
    p = argparse.ArgumentParser()
    p.add_argument('input', help='输入 .xlsx 文件路径')
    p.add_argument('--strict', action='store_true', help='严格模式：任何 error/warn 都失败')
    p.add_argument('--sample', type=int, default=5, help='抽样行数 (默认 5)')
    p.add_argument('--max-issues', type=int, default=50, help='展示的问题最大数 (默认 50)')
    args = p.parse_args()

    print(f'━━━━━━━━━━ 校验 {args.input} ━━━━━━━━━━\n')
    wb = load_workbook(args.input, read_only=True, data_only=True)
    ws = wb.active
    print(f'Sheet: {ws.title}  总行数: {ws.max_row}\n')

    # 1. 表头
    print('── 1. 表头检查 ──')
    header_issues, header = check_header(ws)
    if header_issues:
        for iss in header_issues:
            print(str(iss))
    else:
        print('  ✓ 表头 11 列 OK')
    print()

    # 2. 行级检查
    print('── 2. 行级检查 ──')
    row_issues, stats = check_rows(ws)
    error_issues = [i for i in row_issues if i.level == 'error']
    warn_issues = [i for i in row_issues if i.level == 'warn']

    print(f'\n=== 统计 ===')
    print(f'  错误: {len(error_issues)}')
    print(f'  警告: {len(warn_issues)}')
    print()
    print(f'  题型分布: {dict(stats["type_dist"])}')
    print(f'  题干为空: {stats["empty_stem"]}')
    print(f'  正确答案为空: {stats["empty_ans"]}')
    print(f'  答案格式错: {stats["ans_format_bad"]}')
    print(f'  题型非法: {stats["illegal_type"]}')
    print(f'  题干含题号前缀: {stats["qno_prefix"]}')
    print(f'  题干含章节标签: {stats["section_prefix"]}')
    print(f'  题干疑似多题挤一行: {stats["multi_in_one"]}')
    print(f'  题干过短: {stats["short_stem"]}')
    print(f'  题干仅数字/单位: {stats["digits_only_stem"]}')
    print(f'  题干过长 (>250字): {stats["too_long_stem"]}')

    # 致命错误优先
    if error_issues:
        print(f'\n=== 错误详情 (前 {min(args.max_issues, len(error_issues))} 条) ===')
        for iss in error_issues[:args.max_issues]:
            print(str(iss))
    if warn_issues:
        print(f'\n=== 警告详情 (前 {min(args.max_issues, len(warn_issues))} 条) ===')
        for iss in warn_issues[:args.max_issues]:
            print(str(iss))

    # 3. 重复检查
    print('\n── 3. 重复检查 ──')
    dup_issues, dup_count = check_duplicates(ws)
    if dup_count == 0:
        print('  ✓ 无重复题')
    else:
        print(f'  ⚠ 有 {dup_count} 组重复题（同题干）')
        for iss in dup_issues:
            print(str(iss))

    # 4. 抽样肉眼复核
    print(f'\n── 4. 随机抽样 {args.sample} 行（请肉眼核对）──')
    samples = sample_rows(ws, args.sample)
    for i, row in enumerate(samples, 1):
        print(f'\n  [样本 {i}]')
        print(f'    题干: {(row[0] or "")[:120]}')
        print(f'    题型: {row[1]}')
        opts = [(chr(65+j), str(row[2+j] or "").strip()) for j in range(8)]
        opts_str = ' | '.join(f'{l}:{o[:25]}' for l, o in opts if o)
        print(f'    选项: {opts_str}')
        print(f'    答案: {row[10]}')

    # 总结
    print('\n━━━━━━━━━━ 总结 ━━━━━━━━━━')
    total_errors = len(header_issues) + len(error_issues)
    total_warns = len(warn_issues) + len(dup_issues)
    if total_errors == 0 and total_warns == 0:
        print('🎉 全部通过！可以导入')
        sys.exit(0)
    elif total_errors == 0:
        print(f'⚠ {total_warns} 个警告，建议修复后再导入')
        sys.exit(2 if args.strict else 0)
    else:
        print(f'❌ {total_errors} 个致命错误 + {total_warns} 个警告，必须修复')
        sys.exit(1)


if __name__ == '__main__':
    main()
