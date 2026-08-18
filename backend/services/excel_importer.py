import re
import openpyxl
from sqlalchemy.orm import Session
from models.question_bank import QuestionBank


def normalize_text(text: str) -> str:
    """归一化文本用于匹配：去空格、全角半角转换、去标点、去常见虚词"""
    if not text:
        return ""
    t = text.strip()
    # 全角转半角
    result = []
    for ch in t:
        code = ord(ch)
        if 0xFF01 <= code <= 0xFF5E:
            result.append(chr(code - 0xFEE0))
        elif code == 0x3000:
            result.append(' ')
        else:
            result.append(ch)
    t = ''.join(result)
    # 去空格和标点
    t = re.sub(r'[\s,，。．.、；;：:！!？?（）()【】《》“”‘’\[\]{}]', '', t)
    # 去常见虚词
    t = re.sub(r'[的了是吗呢啊吧呀嘛哪着过会被从把让给对向到在]', '', t)
    return t.lower()


def import_from_excel(file_path: str, db: Session) -> dict:
    """从Excel文件导入题库"""
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    # 找表头行（第3行是表头）
    header_row = 3
    headers = {}
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=header_row, column=col).value
        if val:
            headers[col] = str(val).strip()

    imported = 0
    skipped = 0
    errors = []

    for row in range(header_row + 1, ws.max_row + 1):
        try:
            row_data = {}
            for col, name in headers.items():
                cell_val = ws.cell(row=row, column=col).value
                row_data[name] = str(cell_val).strip() if cell_val is not None else ""

            stem = row_data.get("题干", "")
            if not stem:
                skipped += 1
                continue

            q = QuestionBank(
                question_type=row_data.get("题型", ""),
                stem=stem,
                stem_normalized=normalize_text(stem),
                question_image=row_data.get("题目图片", ""),
                option_a=row_data.get("选项A", ""),
                option_a_image=row_data.get("选项A图片", ""),
                option_b=row_data.get("选项B", ""),
                option_b_image=row_data.get("选项B图片", ""),
                option_c=row_data.get("选项C", ""),
                option_c_image=row_data.get("选项C图片", ""),
                option_d=row_data.get("选项D", ""),
                option_d_image=row_data.get("选项D图片", ""),
                option_e=row_data.get("选项E", ""),
                option_e_image=row_data.get("选项E图片", ""),
                option_f=row_data.get("选项F", ""),
                option_f_image=row_data.get("选项F图片", ""),
                option_g=row_data.get("选项G", ""),
                option_g_image=row_data.get("选项G图片", ""),
                option_h=row_data.get("选项H", ""),
                option_h_image=row_data.get("选项H图片", ""),
                correct_answer=row_data.get("正确答案", ""),
                explanation=row_data.get("解析", ""),
                explanation_image=row_data.get("解析图片", ""),
                subject=row_data.get("科目", ""),
                chapter=row_data.get("章节", ""),
            )
            db.add(q)
            imported += 1
        except Exception as e:
            errors.append(f"第{row}行导入失败: {str(e)}")
            skipped += 1

    db.commit()
    return {"imported": imported, "skipped": skipped, "errors": errors}


def clear_bank(db: Session) -> int:
    count = db.query(QuestionBank).count()
    db.query(QuestionBank).delete()
    db.commit()
    return count


def get_stats(db: Session) -> list:
    from sqlalchemy import func as sqlfunc
    rows = (
        db.query(
            QuestionBank.subject,
            QuestionBank.question_type,
            sqlfunc.count(QuestionBank.id)
        )
        .group_by(QuestionBank.subject, QuestionBank.question_type)
        .order_by(QuestionBank.subject, QuestionBank.question_type)
        .all()
    )
    stats = {}
    for subject, qtype, count in rows:
        s = subject or "未分类"
        if s not in stats:
            stats[s] = {"subject": s, "total": 0, "types": {}}
        stats[s]["total"] += count
        stats[s]["types"][qtype or "未知"] = count
    return list(stats.values())
